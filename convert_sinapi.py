#!/usr/bin/env python3
"""
Converte a planilha oficial SINAPI (mensal) em sinapi_data.json, o arquivo de
dados usado pelo Planejador de Obras (index.html) — composições, mão de obra,
materiais, equipamentos E os preços (R$) usados na aba Orçamento.

IMPORTANTE: a partir desta versão, o script precisa da planilha SINAPI
"COMPLETA" (a que tem as abas Menu, Busca, ISD, ICD, ISE, CSD, CCD, CSE,
Analítico, Analítico com Custo) — não só a aba "Analítico" isolada. É o
arquivo que o site da Caixa/SINAPI disponibiliza todo mês como
"SINAPI Referência de Preços e Custos".

Os preços usados são os da coluna do Piauí (PI), sem desoneração (ISD/CSD) —
o padrão mais comum para construtoras. Se quiser outro estado ou a versão
"com desoneração", veja a função extract_prices() abaixo e troque o índice
de coluna.

USO (todo mês, quando sai a nova base SINAPI):
    python3 convert_sinapi.py "SINAPI_Referencia_2026_07.xlsx"

Isso sobrescreve sinapi_data.json na mesma pasta. Depois é só publicar os
arquivos junto com o resto do site (GitHub → Vercel).

Requer: pip install openpyxl --break-system-packages
"""
import sys
import re
import json
import os

# Ordem das siglas de estado como aparecem nas planilhas SINAPI (ISD/CSD).
# Mude aqui se um dia precisar trocar de estado.
ESTADO_ALVO = "PI"


def find_state_col_isd(ws):
    """ISD/ICD/ISE: uma coluna por estado, a partir da linha de cabeçalho."""
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if row and row[0] == "Classificação":
            for i, v in enumerate(row):
                if v == ESTADO_ALVO:
                    return i
    return None


def find_state_col_csd(ws):
    """CSD/CCD/CSE: cada estado ocupa 2 colunas (Custo(R$) + %AS)."""
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        if row and ESTADO_ALVO in row:
            return row.index(ESTADO_ALVO)
    return None


def extract_prices(wb_formulas, wb_values):
    """Retorna (comp_price, insumo_price): dict codigo -> preço R$ no Piauí."""
    comp_price = {}
    if "CSD" in wb_values.sheetnames:
        ws_f = wb_formulas["CSD"]
        ws_v = wb_values["CSD"]
        col = find_state_col_csd(ws_v)
        if col is not None:
            for rf, rv in zip(ws_f.iter_rows(min_row=11, values_only=True),
                               ws_v.iter_rows(min_row=11, values_only=True)):
                codf = rf[1] if len(rf) > 1 else None
                if codf is None:
                    continue
                nums = re.findall(r"\d+", str(codf))
                if not nums:
                    continue
                code = int(nums[-1])
                val = rv[col] if len(rv) > col else None
                if isinstance(val, (int, float)):
                    comp_price[code] = round(float(val), 4)

    insumo_price = {}
    if "ISD" in wb_values.sheetnames:
        ws = wb_values["ISD"]
        col = find_state_col_isd(ws)
        if col is not None:
            for row in ws.iter_rows(min_row=11, values_only=True):
                code = row[1] if len(row) > 1 else None
                if code is None:
                    continue
                val = row[col] if len(row) > col else None
                if isinstance(val, (int, float)):
                    insumo_price[int(code)] = round(float(val), 4)

    return comp_price, insumo_price


def guess_referencia(xlsx_path):
    """Tenta achar 'AAAA_MM' ou 'AAAA-MM' no nome do arquivo (ex: SINAPI_Referencia_2026_07.xlsx)
    e devolve como 'AAAA/MM'. Se não achar, devolve None (o app mostra 'não identificada')."""
    name = os.path.basename(xlsx_path)
    m = re.search(r"(20\d{2})[_\-](\d{2})\b", name)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    return None


def convert(xlsx_path, out_path=None):
    import openpyxl

    if out_path is None:
        out_path = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)) or ".", "sinapi_data.json")

    print(f"Lendo {xlsx_path} ...")
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    has_full = "CSD" in wb_values.sheetnames or "ISD" in wb_values.sheetnames
    if has_full:
        wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=True)
        print("Planilha completa detectada — preços (Piauí) serão incluídos.")
    else:
        wb_formulas = None
        print("AVISO: só a aba 'Analítico' foi encontrada — sem abas de preço (ISD/CSD).")
        print("       O arquivo vai sair sem preços (aba Orçamento não vai auto-preencher R$).")
        print("       Baixe a planilha SINAPI completa (não só a 'Analítico') para ter preços.")

    ws = wb_values["Analítico"] if "Analítico" in wb_values.sheetnames else wb_values[wb_values.sheetnames[0]]

    items = {}      # codigo -> [descricao, unidade]  (preço é anexado depois)
    children = {}    # codigo_pai -> [[tipo(0=insumo,1=composicao), codigo_filho, coeficiente], ...]
    tops = []         # [grupo, codigo, descricao, unidade]  (composições selecionáveis)

    n_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        grupo, cod_comp, tipo, cod_item, desc, un, coef, situacao = (list(row) + [None] * 8)[:8]
        if cod_comp is None:
            continue
        n_rows += 1
        if tipo is None:
            tops.append([grupo, cod_comp, desc, un])
            items.setdefault(cod_comp, [desc, un])
            continue
        items.setdefault(cod_item, [desc, un])
        tflag = 0 if tipo == "INSUMO" else 1
        children.setdefault(cod_comp, []).append([tflag, cod_item, coef])

    n_priced = 0
    if has_full:
        comp_price, insumo_price = extract_prices(wb_formulas, wb_values)
        for code, val in items.items():
            price = insumo_price.get(code)
            if price is None:
                price = comp_price.get(code)
            val.append(price)
            if price is not None:
                n_priced += 1
    else:
        for val in items.values():
            val.append(None)

    data = {"items": items, "children": children, "tops": tops, "meta": {"referencia": guess_referencia(xlsx_path)}}

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    size_mb = os.path.getsize(out_path) / 1e6
    print(f"OK: {len(tops)} composições, {len(items)} itens únicos ({n_priced} com preço), {n_rows} linhas.")
    print(f"Arquivo gerado: {out_path} ({size_mb:.2f} MB)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python3 convert_sinapi.py <planilha_sinapi_completa.xlsx>")
        sys.exit(1)
    convert(sys.argv[1])
